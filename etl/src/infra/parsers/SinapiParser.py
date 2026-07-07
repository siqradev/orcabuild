#src/infra/parsers/SinapiParser.py
import zipfile
import openpyxl
import json
import sys
import os
from io import BytesIO

def parse_sinapi_zip(zip_path, target_state):
    if not os.path.exists(zip_path):
        return {"error": f"Arquivo não encontrado: {zip_path}"}

    items = []
    
    try:
        with zipfile.ZipFile(zip_path, 'r') as z:
            # Filtramos para pegar o Excel que contém os Preços de Insumos
            excel_files = [f for f in z.namelist() if f.endswith('.xlsx') and not f.startswith('__')]
            
            if not excel_files:
                return {"error": "Nenhum arquivo .xlsx encontrado dentro do ZIP."}

            with z.open(excel_files[0]) as f:
                content = f.read()
                wb = openpyxl.load_workbook(filename=BytesIO(content), data_only=True)
                
                # ERRO PREVENIDO: Selecionamos a aba ISD (Insumos Sem Desoneração) explicitamente
                # Os seus arquivos mostram que abas como 'Menu' ou 'Analítico' viriam primeiro se usarmos wb.active
                if 'ISD' in wb.sheetnames:
                    sheet = wb['ISD']
                else:
                    sheet = wb.active # Fallback caso o nome mude

                # MAPEAMENTO REAL (Baseado nos seus arquivos enviados):
                # Coluna A: Classe/Tipo
                # Coluna B: Código do Insumo
                # Coluna C: Descrição do Insumo
                # Coluna D: Unidade
                # Coluna E: Origem do Preço
                # Colunas F em diante: Estados (AC, AL, AM, AP, BA, CE...)
                
                # Precisamos achar a coluna do estado (CE por exemplo)
                header_row = list(sheet.iter_rows(min_row=6, max_row=6, values_only=True))[0]
                try:
                    state_col_index = header_row.index(target_state)
                except (ValueError, IndexError):
                    # Se não achar o estado no cabeçalho, tenta um padrão (CE costuma ser a coluna 10 / index 10)
                    state_col_index = 10 

                for row in sheet.iter_rows(min_row=7, values_only=True):
                    # Validação de segurança: o código deve ser preenchido
                    raw_code = row[1] # Coluna B
                    if not raw_code:
                        continue
                    
                    code = str(raw_code).strip()
                    if not code.isdigit(): # Pula linhas de títulos internos (ex: "MATERIAIS")
                        continue

                    try:
                        price_val = row[state_col_index]
                        # O openpyxl com data_only=True já traz o número, 
                        # mas se vier string "15,30", o float() falharia sem o replace.
                        if isinstance(price_val, str):
                            price = float(price_val.replace('.', '').replace(',', '.'))
                        else:
                            price = float(price_val) if price_val is not None else 0.0
                    except:
                        price = 0.0

                    items.append({
                        "code": code,
                        "description": str(row[2]).strip(), # Coluna C
                        "unit": str(row[3]).strip() if row[3] else "", # Coluna D
                        "price": price
                    })

        return items

    except Exception as e:
        return {"error": str(e)}

if __name__ == "__main__":
    # Agora recebemos o caminho e o estado: python3 parser.py arquivo.zip CE
    file_to_parse = sys.argv[1]
    state = sys.argv[2] if len(sys.argv) > 2 else "CE"
    
    result = parse_sinapi_zip(file_to_parse, state)
    print(json.dumps(result))